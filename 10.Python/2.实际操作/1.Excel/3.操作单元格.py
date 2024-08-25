from openpyxl import Workbook

# 1. 基本方法
# wb = Workbook()
# ws = wb.active

# #修改单元格 方法1
# ws["a6"] = "第a列第6行"

# #修改单元格 方法2
# cell = ws.cell(6, 2, "第6行第二列")
# cell.value = 99
# print(cell.value)
# print(cell.coordinate)
# print(cell.row)
# print(cell.column)
# print(cell.column_letter)


# wb.save("./VSCodeProjects/test2.xlsx")

# 2.二维遍历
wb = Workbook()
ws = wb.active

x = 1
for i in range(1,11):
    for j in range(1,6):
        ws.cell(i,j,x)
        x +=1

wb.save("./VSCodeProjects/test2.xlsx")

# 3.以列或行为单位遍历 元组套元组
print(ws["a:c"])
print(ws[1:5])
print(ws["a1:c4"])
print(ws["b"])
#拆分元组
for cells in ws["a1:c4"]:
    for cell in cells:
        print(cell.value)