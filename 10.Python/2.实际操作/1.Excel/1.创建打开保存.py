from openpyxl import Workbook, load_workbook

# # 1.新建并保存一个工作簿
# wb = Workbook()
# # active有一个装饰器@property,使得这个方法可以当成一个属性来调用
# # 会返回一个当前已经打开的工作表,获取第一个sheet
# ws = wb.active
# # 打印这个工作簿的名字
# print(ws.title)
# # 保存到哪里?
# wb.save("/Users/peihanggu/GithubLocal/Python/VSCodeProjects/test1.xlsx")

# 2.打开刚才的工作簿
wb = load_workbook("/Users/peihanggu/GithubLocal/Python/VSCodeProjects/test1.xlsx")
ws = wb.active
print(ws.title)

