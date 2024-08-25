from openpyxl import Workbook, load_workbook

wb = Workbook()
ws1 = wb.active
print(ws1.title)
#获得所有工作表,也是装饰器
print(wb.sheetnames)

# 创建工作簿,工作簿名,位置从0开始
wb.create_sheet("Sheet2",1)
wb.create_sheet("Sheet3",2)
print(wb.sheetnames)

#字典的方式取得某一个sheet
ws3 = wb["Sheet3"]
print(ws3.title)

#更改工作表顺序, offset 0是原地,-1是左
wb.move_sheet(ws3, -1)
print(wb.sheetnames)

#删除工作表
del wb["Sheet3"]
print(wb.sheetnames)

#复制工作表?
